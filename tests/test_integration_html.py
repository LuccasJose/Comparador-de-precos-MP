import contextlib
import difflib
import functools
import http.server
import json
import os
import pathlib
import re
import shutil
import sys
import threading
import types
import unittest
from unittest.mock import patch

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "src")))

PROJECT_ROOT = pathlib.Path(__file__).resolve().parents[1]
URL_MEU = "https://livemenu.app/menu/64e37fc6a53f6d0056007d55"
URL_CONC = "https://pedido.anota.ai/loja/caseratto-passeio-das-guas-1"
PRICE_RE = re.compile(r"(?:R\$\s*)?([0-9]{1,6}(?:[\.,:][0-9]{2}))")
HAS_LETTER_RE = re.compile(r"[A-Za-zÀ-ÿ]")


def _price_to_float(raw):
    raw = raw.replace("R$", "").replace(":", ".").strip()
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    return round(float(raw), 2)


def _extract_items(text, limit=20):
    lines = [line.strip(" -•\t") for line in str(text).splitlines() if line.strip()]
    items, seen = [], set()
    for idx, line in enumerate(lines):
        matches = list(PRICE_RE.finditer(line))
        if not matches:
            continue
        price_raw = matches[-1].group(0)
        name = PRICE_RE.sub("", line).strip(" -•.\t")
        if (not name or not HAS_LETTER_RE.search(name)) and idx > 0:
            prev = lines[idx - 1].strip()
            if HAS_LETTER_RE.search(prev) and not PRICE_RE.search(prev):
                name = prev
        if not name or not HAS_LETTER_RE.search(name):
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append({"n": name[:120], "p": _price_to_float(price_raw)})
        if len(items) >= limit:
            break
    return items


def _mock_generate_menu_json_with_cache(_prompt, txt_meu, txt_conc, **_kwargs):
    meu = _extract_items(txt_meu)
    conc = _extract_items(txt_conc)
    if not meu or not conc:
        raise AssertionError("Mock do Gemini não conseguiu extrair itens suficientes do OCR otimizado.")
    return {"meu": meu, "conc": conc}, {"fonte": "cache", "total_tokens": 0}


def _install_import_fallbacks():
    if "dotenv" not in sys.modules:
        dotenv = types.ModuleType("dotenv")
        dotenv.load_dotenv = lambda *args, **kwargs: False
        dotenv.find_dotenv = lambda *args, **kwargs: ""
        sys.modules["dotenv"] = dotenv
    if "thefuzz" not in sys.modules:
        def _ratio(a, b):
            return int(difflib.SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio() * 100)
        def _extract_one(query, choices, scorer=None):
            scorer = scorer or _ratio
            ranked = sorted(((choice, scorer(query, choice)) for choice in choices), key=lambda item: item[1], reverse=True)
            return ranked[0] if ranked else None
        thefuzz = types.ModuleType("thefuzz")
        thefuzz.fuzz = types.SimpleNamespace(token_sort_ratio=_ratio)
        thefuzz.process = types.SimpleNamespace(extractOne=_extract_one)
        sys.modules["thefuzz"] = thefuzz
    if "pandas" not in sys.modules:
        pandas = types.ModuleType("pandas")
        pandas.DataFrame = lambda rows: rows
        sys.modules["pandas"] = pandas


def _import_main_module():
    try:
        import main
        return main
    except ModuleNotFoundError as exc:
        if exc.name not in {"dotenv", "thefuzz", "pandas"}:
            raise
        _install_import_fallbacks()
        import importlib
        return importlib.import_module("main")


def _write_xlsx(rows, path):
    from openpyxl import Workbook

    headers = list(rows[0].keys()) if rows else ["Concorrente", "Meu Item", "Preco Conc", "Meu Preco", "Diferenca", "Status"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    wb.save(path)
    wb.close()


class _LocalHttpServer:
    def __init__(self, directory):
        self.directory = str(directory)
        self.server = None
        self.thread = None

    def __enter__(self):
        handler = functools.partial(http.server.SimpleHTTPRequestHandler, directory=self.directory)
        self.server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), handler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        host, port = self.server.server_address
        return f"http://{host}:{port}"

    def __exit__(self, exc_type, exc, tb):
        if self.server:
            self.server.shutdown()
            self.server.server_close()
        if self.thread:
            self.thread.join(timeout=5)


class TestIntegrationHtml(unittest.TestCase):
    @unittest.skipUnless(os.getenv("RUN_HTML_E2E") == "1", "Defina RUN_HTML_E2E=1 para executar o teste E2E com scraping real.")
    def test_full_flow_generates_files_and_renders_index(self):
        from openpyxl import load_workbook

        main = _import_main_module()

        dados_json = PROJECT_ROOT / "dados.json"
        xlsx_path = PROJECT_ROOT / "comparativo_precos.xlsx"
        backups = {}
        for path in (dados_json, xlsx_path):
            if path.exists():
                backup = path.with_name(path.name + ".e2e.bak")
                shutil.copy2(path, backup)
                backups[path] = backup
                path.unlink()
        try:
            patcher = contextlib.nullcontext()
            if not os.getenv("GEMINI_API_KEY"):
                patcher = patch.object(main, "generate_menu_json_with_cache", side_effect=_mock_generate_menu_json_with_cache)
            with patcher:
                relatorio, dados = main.comparar(URL_MEU, URL_CONC, timeout_seconds=45, headless=True)
            dados_json.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
            _write_xlsx(relatorio, xlsx_path)

            payload = json.loads(dados_json.read_text(encoding="utf-8"))
            self.assertTrue(payload["meu"] and payload["conc"])
            self.assertTrue(all(item.get("n") and isinstance(item.get("p"), (int, float)) for item in payload["meu"] + payload["conc"]))
            self.assertTrue(xlsx_path.exists() and xlsx_path.stat().st_size > 0)
            wb = load_workbook(xlsx_path, read_only=True)
            self.assertGreaterEqual(len(wb.sheetnames), 1)
            wb.close()

            from playwright.sync_api import sync_playwright

            with _LocalHttpServer(PROJECT_ROOT) as base_url, sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                page = browser.new_page()
                page.goto(f"{base_url}/index.html", wait_until="networkidle")
                # Wait for dynamic dados.json to load and render the dashboard
                page.wait_for_function(
                    "() => document.querySelectorAll('.comp-row:not(.header-row)').length > 0",
                    timeout=10000,
                )
                # KPI cards should be rendered
                kpi_cards = page.locator(".kpi-card").count()
                self.assertGreaterEqual(kpi_cards, 1, "KPI cards should be rendered")
                # Comparison rows should exist (items from dados.json)
                comp_rows = page.locator(".comp-row:not(.header-row)").count()
                self.assertGreater(comp_rows, 0, "Comparison rows should exist")
                # Verify no dish names start with "Acompanha" (OCR fix)
                all_names = page.locator(".item-name").all_text_contents()
                for name in all_names:
                    name_stripped = name.strip().split("\n")[0].strip()
                    self.assertFalse(
                        name_stripped.lower().startswith("acompanha"),
                        f"Dish name should not start with 'Acompanha': {name_stripped}",
                    )
                browser.close()
        finally:
            for path in (dados_json, xlsx_path):
                if path.exists():
                    path.unlink()
            for path, backup in backups.items():
                shutil.move(str(backup), str(path))


if __name__ == "__main__":
    unittest.main()

